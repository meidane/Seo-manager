"""ویوهای حسابداری (پایه) — داشبورد، بانک‌ها، بابت‌ها، تراکنش‌ها، ورود اکسل، حقوق."""
import json

from datetime import date

from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.paginator import Paginator
from django.db.models import Max, Q, Sum
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from django.views.decorators.http import require_http_methods
from django.views.generic import TemplateView

from core.daterange import (PRESET_LABELS, PRESETS, DateRangeMixin,
                            _resolve_preset)
from core.jalali import format_jalali, parse_jalali
from projects.models import Project

from .access import FinancePermMixin, require_finance
from .models import (BankAccount, Category, Invoice, InvoiceLine, Payroll,
                     PayrollItem, Transaction)
from .utils import parse_amount, parse_excel_date


def _pj(value):
    """تاریخ شمسی → میلادی؛ خالی/نامعتبر → None."""
    if not value:
        return None
    try:
        return parse_jalali(value)
    except (ValueError, TypeError):
        return None


def _body(request):
    try:
        return json.loads(request.body or '{}')
    except json.JSONDecodeError:
        return {}


# ── صفحات ────────────────────────────────────────────────────────────────

class FinanceDashboardView(LoginRequiredMixin, FinancePermMixin, DateRangeMixin, TemplateView):
    template_name = 'finance/dashboard.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        start, end = self.get_range(self.request)
        ctx.update(self.range_context())
        tx = Transaction.objects.filter(date__range=(start, end))

        income = tx.aggregate(s=Sum('deposit'))['s'] or 0
        expense = tx.aggregate(s=Sum('withdrawal'))['s'] or 0
        banks = list(BankAccount.objects.filter(is_active=True))
        bank_total = sum(b.balance for b in banks)

        # تفکیک بابت: درآمد/هزینه در هر دسته
        cats = []
        for c in Category.objects.all():
            ct = tx.filter(category=c)
            cats.append({
                'name': c.name, 'color': c.color,
                'income': ct.aggregate(s=Sum('deposit'))['s'] or 0,
                'expense': ct.aggregate(s=Sum('withdrawal'))['s'] or 0,
            })

        # حقوق: تعهد کل و پرداختی (کل، مستقل از بازه)
        pay_total = sum(p.total for p in Payroll.objects.all())
        pay_paid = Payroll.objects.aggregate(s=Sum('paid_amount'))['s'] or 0

        from .alerts import compute_alerts
        ctx.update({
            'income': income, 'expense': expense, 'net': income - expense,
            'banks': banks, 'bank_total': bank_total, 'cats': cats,
            'pay_total': pay_total, 'pay_paid': pay_paid, 'pay_remaining': pay_total - pay_paid,
            'unassigned': tx.filter(project__isnull=True, category__isnull=True).count(),
            'fin_alerts': compute_alerts(),
            'page_title': 'حسابداری',
        })
        return ctx


def _optional_range(g):
    """بازه‌ی تاریخِ اختیاری برای تراکنش‌ها — پیش‌فرض بدون فیلتر (همه).
    فقط وقتی کاربر صریح `from/to` یا `range` بدهد اعمال می‌شود.
    برمی‌گرداند (start, end, label) یا (None, None, '')."""
    if g.get('from') and g.get('to'):
        try:
            s, e = parse_jalali(g['from']), parse_jalali(g['to'])
            return s, e, f'{format_jalali(s)} تا {format_jalali(e)}'
        except (ValueError, TypeError):
            pass
    key = g.get('range')
    if key and (key in PRESETS or key in ('this_month', 'last_month')):
        s, e = _resolve_preset(key, date.today())
        return s, e, PRESET_LABELS.get(key, '')
    return None, None, ''


class TransactionListView(LoginRequiredMixin, FinancePermMixin, TemplateView):
    template_name = 'finance/transactions.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        g = self.request.GET
        qs = Transaction.objects.select_related('bank_account', 'project', 'category')

        # بازه‌ی تاریخ اختیاری (پیش‌فرض: همه‌ی تراکنش‌ها)
        start, end, range_label = _optional_range(g)
        if start and end:
            qs = qs.filter(date__range=(start, end))

        if g.get('bank'):
            qs = qs.filter(bank_account_id=g['bank'])
        if g.get('project'):
            qs = qs.filter(project_id=g['project'])
        # بابت: چندانتخابی
        cat_ids = [c for c in g.getlist('category') if c]
        if cat_ids:
            qs = qs.filter(category_id__in=cat_ids)
        if g.get('unassigned') == '1':
            qs = qs.filter(project__isnull=True, category__isnull=True)
        # جستجو روی شرح سند + توضیحات
        q = (g.get('q') or '').strip()
        if q:
            qs = qs.filter(Q(description__icontains=q) | Q(note__icontains=q))

        paginator = Paginator(qs, 100)
        page_obj = paginator.get_page(g.get('page'))

        # پارامترهای فیلتر برای لینک‌های صفحه‌بندی (بدون page)
        params = g.copy()
        params.pop('page', None)
        ctx['qs_params'] = params.urlencode()

        ctx['page_obj'] = page_obj
        ctx['transactions'] = page_obj.object_list
        ctx['total_count'] = paginator.count
        ctx['banks'] = BankAccount.objects.filter(is_active=True)
        ctx['projects'] = Project.objects.filter(status=Project.ACTIVE)
        ctx['categories'] = Category.objects.all()
        ctx['selected_categories'] = cat_ids
        ctx['range_label'] = range_label
        ctx['q'] = q
        ctx['filters'] = g
        ctx['page_title'] = 'تراکنش‌ها'
        return ctx


class BankListView(LoginRequiredMixin, FinancePermMixin, TemplateView):
    template_name = 'finance/banks.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['banks'] = BankAccount.objects.all()
        ctx['categories'] = Category.objects.all()
        ctx['page_title'] = 'بانک‌ها'
        return ctx


class ImportView(LoginRequiredMixin, FinancePermMixin, TemplateView):
    template_name = 'finance/import.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['banks'] = BankAccount.objects.filter(is_active=True)
        ctx['page_title'] = 'ورود اکسل'
        return ctx


class PayrollListView(LoginRequiredMixin, FinancePermMixin, TemplateView):
    template_name = 'finance/payroll.html'

    def get_context_data(self, **kwargs):
        from colleagues.models import Colleague
        from core.jalali import MONTH_NAMES

        from .balances import salary_balances
        ctx = super().get_context_data(**kwargs)
        payrolls = list(Payroll.objects.select_related('colleague').prefetch_related('items'))
        cids = {p.colleague_id for p in payrolls}

        # مانده‌ی «کل حساب با همکار» (منبع واحد: balances.salary_balances)
        balances = salary_balances(cids)
        # نگاشتِ همکار → id بابتِ حقوقِ او، برای لینکِ مانده به تبِ گزارش
        salary_cat = {c.colleague_id: c.id for c in Category.objects.filter(colleague_id__in=cids)}

        ctx['payrolls'] = payrolls
        ctx['balances'] = balances
        ctx['salary_cat'] = salary_cat
        ctx['colleagues'] = Colleague.objects.filter(status=Colleague.ACTIVE)
        ctx['months'] = list(enumerate(MONTH_NAMES, start=1))  # [(1,'فروردین'),...]
        ctx['page_title'] = 'حقوق'
        return ctx


class InvoiceListView(LoginRequiredMixin, FinancePermMixin, DateRangeMixin, TemplateView):
    template_name = 'finance/invoices.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        start, end = self.get_range(self.request)
        ctx.update(self.range_context())
        qs = (Invoice.objects.select_related('project')
              .prefetch_related('lines')
              .filter(issue_date__range=(start, end)))
        if self.request.GET.get('project'):
            qs = qs.filter(project_id=self.request.GET['project'])
        invoices = list(qs)
        # مانده‌ی «گردش حساب» هر پروژه (کلِ تاریخ) برای ستونِ مانده — منبع واحد balances
        from .balances import project_balances
        ctx['project_bal'] = project_balances({inv.project_id for inv in invoices})
        ctx['invoices'] = invoices
        ctx['projects'] = Project.objects.filter(status=Project.ACTIVE)
        ctx['filters'] = self.request.GET
        ctx['page_title'] = 'فاکتورها'
        return ctx


class InvoiceFormView(LoginRequiredMixin, FinancePermMixin, TemplateView):
    """صفحه‌ی ساخت/ویرایشِ فاکتور (فرمِ کامل با ردیف‌های پویا)."""

    template_name = 'finance/invoice_form.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        pk = kwargs.get('pk')
        invoice = None
        if pk:
            invoice = get_object_or_404(
                Invoice.objects.prefetch_related('lines__category'), pk=pk)
        ctx['invoice'] = invoice
        ctx['lines'] = invoice.lines.all() if invoice else []
        # پیش‌نمایشِ شماره‌ی بعدی برای فاکتورِ جدید
        if not invoice:
            last = Invoice.objects.aggregate(m=Max('number'))['m']
            ctx['next_number'] = (last or 0) + 1
        ctx['projects'] = Project.objects.filter(status=Project.ACTIVE)
        ctx['categories'] = Category.objects.all()
        ctx['page_title'] = f'فاکتور #{invoice.number}' if invoice else 'فاکتور جدید'
        return ctx


class LedgerView(LoginRequiredMixin, FinancePermMixin, DateRangeMixin, TemplateView):
    """گردشِ حساب (مثلِ صورت‌حسابِ بانکی) — فیلترِ **یا** پروژه **یا** بابت (نه هر دو، نه هیچ‌کدام).

    - پروژه: فاکتورهای پروژه = برداشت؛ تراکنش‌های پروژه (واریز/برداشت) = واریز/برداشت.
    - بابت: تراکنش‌های آن بابت (واریز/برداشت)؛ اگر بابتِ حقوقِ یک همکار بود، حقوق‌های او
      در بازه (تاریخِ اولِ ماهِ ثبت) هم به‌عنوان واریز اضافه می‌شود.
    مانده به‌صورت تجمعی (Σواریز − Σبرداشت) در هر ردیف، و جمعِ نهایی در ردیفِ آخر.
    """

    template_name = 'finance/ledger.html'

    def get_context_data(self, **kwargs):
        from core.jalali import format_jalali, j2g
        ctx = super().get_context_data(**kwargs)
        start, end = self.get_range(self.request)
        ctx.update(self.range_context())
        g = self.request.GET

        project_id = g.get('project') or ''
        category_id = g.get('category') or ''
        bank_id = g.get('bank') or ''

        ctx['banks'] = BankAccount.objects.filter(is_active=True)
        ctx['projects'] = Project.objects.filter(status=Project.ACTIVE)
        ctx['categories'] = Category.objects.all()
        ctx['filters'] = g
        ctx['page_title'] = 'گزارش (گردش حساب)'

        # شرطِ طلایی: دقیقاً یکی از پروژه/بابت انتخاب شده باشد
        if bool(project_id) == bool(category_id):
            ctx['rows'] = None
            ctx['invalid_filter'] = True
            return ctx

        rows = []

        if project_id:
            ctx['mode'] = 'project'
            ctx['selected_project'] = Project.objects.filter(id=project_id).first()
            # فاکتورهای پروژه → برداشت
            for inv in Invoice.objects.filter(project_id=project_id, issue_date__range=(start, end)):
                rows.append({
                    'date': inv.issue_date,
                    'title': f'فاکتور #{inv.number}' + (f' — {inv.description}' if inv.description else ''),
                    'kind': 'invoice', 'ref_id': inv.id,
                    'deposit': 0, 'withdrawal': inv.grand_total, 'bank': '',
                })
            # تراکنش‌های پروژه → واریز/برداشت
            tx = Transaction.objects.select_related('bank_account').filter(
                project_id=project_id, date__range=(start, end))
            if bank_id:
                tx = tx.filter(bank_account_id=bank_id)
            for t in tx:
                rows.append({
                    'date': t.date, 'title': t.description or '—',
                    'kind': 'tx', 'ref_id': t.id,
                    'deposit': t.deposit or 0, 'withdrawal': t.withdrawal or 0,
                    'bank': t.bank_account.name if t.bank_account_id else '',
                })

        else:  # category_id
            ctx['mode'] = 'category'
            cat = Category.objects.filter(id=category_id).select_related('colleague').first()
            ctx['selected_category'] = cat
            # تراکنش‌های این بابت → واریز/برداشت
            tx = Transaction.objects.select_related('bank_account').filter(
                category_id=category_id, date__range=(start, end))
            if bank_id:
                tx = tx.filter(bank_account_id=bank_id)
            for t in tx:
                rows.append({
                    'date': t.date, 'title': t.description or '—',
                    'kind': 'tx', 'ref_id': t.id,
                    'deposit': t.deposit or 0, 'withdrawal': t.withdrawal or 0,
                    'bank': t.bank_account.name if t.bank_account_id else '',
                })
            # بابتِ حقوق → حقوق‌های همکار در بازه (تاریخ = اولِ ماهِ ثبت) به‌عنوان واریز
            if cat and cat.colleague_id:
                for p in Payroll.objects.filter(colleague_id=cat.colleague_id).prefetch_related('items'):
                    try:
                        d = j2g(p.year, p.month, 1)
                    except (ValueError, TypeError):
                        continue
                    if start <= d <= end:
                        rows.append({
                            'date': d, 'title': f'حقوق {p.month_name} {p.year}',
                            'kind': 'payroll', 'ref_id': p.id,
                            'deposit': p.total, 'withdrawal': 0, 'bank': '',
                        })

        # مرتب‌سازی بر اساس تاریخ (پایدار) + مانده‌ی تجمعی
        rows.sort(key=lambda r: r['date'])
        bal = 0
        tot_d = tot_w = 0
        for r in rows:
            bal += r['deposit'] - r['withdrawal']
            r['balance'] = bal
            r['date_fa'] = format_jalali(r['date'])
            tot_d += r['deposit']
            tot_w += r['withdrawal']

        ctx['rows'] = rows
        ctx['total_deposit'] = tot_d
        ctx['total_withdrawal'] = tot_w
        ctx['final_balance'] = bal

        # بنرِ هشدارِ متنی برای موردِ در حالِ مشاهده (بر پایه‌ی مانده‌ی کلِ تاریخ، نه فقط بازه)
        from .balances import project_balance, salary_balance
        ctx['ledger_alert'] = None
        if project_id:
            pb = project_balance(project_id)
            if pb > 0:
                ctx['ledger_alert'] = 'مانده‌ی این پروژه مثبت است (واریزی بیش از فاکتورها) — شاید فاکتوری ثبت نشده باشد.'
        elif ctx.get('selected_category') and ctx['selected_category'].colleague_id:
            sb = salary_balance(ctx['selected_category'].colleague_id)
            if sb < 0:
                ctx['ledger_alert'] = 'پرداختِ حقوق بیش از تعهد است (اضافه‌پرداخت).'
        return ctx


# ── API: بانک ─────────────────────────────────────────────────────────────

@login_required
@require_finance
@require_http_methods(['POST'])
def bank_create(request):
    d = _body(request)
    if not d.get('name'):
        return JsonResponse({'detail': 'نام لازم است'}, status=400)
    b = BankAccount.objects.create(
        name=d['name'], bank=d.get('bank', ''), color=d.get('color', '#4183F2'),
        card_number=d.get('card_number', ''), initial_balance=parse_amount(d.get('initial_balance', 0)),
        created_by=request.user)
    return JsonResponse({'id': b.id}, status=201)


@login_required
@require_finance
@require_http_methods(['PATCH', 'DELETE'])
def bank_edit(request, pk):
    b = get_object_or_404(BankAccount, pk=pk)
    if request.method == 'DELETE':
        b.delete()
        return JsonResponse({'ok': True})
    d = _body(request)
    for f in ('name', 'bank', 'color', 'card_number'):
        if f in d:
            setattr(b, f, d[f])
    if 'initial_balance' in d:
        b.initial_balance = parse_amount(d['initial_balance'])
    if 'is_active' in d:
        b.is_active = bool(d['is_active'])
    b.save()
    return JsonResponse({'ok': True})


# ── API: بابت ─────────────────────────────────────────────────────────────

@login_required
@require_finance
@require_http_methods(['POST'])
def category_create(request):
    d = _body(request)
    name = (d.get('name') or '').strip()
    if not name:
        return JsonResponse({'detail': 'عنوان لازم است'}, status=400)
    if Category.objects.filter(name=name).exists():
        return JsonResponse({'detail': 'تکراری است'}, status=400)
    c = Category.objects.create(name=name, color=d.get('color', '#8FA0B8'),
                                is_salary=bool(d.get('is_salary')), order=Category.objects.count())
    return JsonResponse({'id': c.id, 'name': c.name}, status=201)


@login_required
@require_finance
@require_http_methods(['DELETE'])
def category_delete(request, pk):
    get_object_or_404(Category, pk=pk).delete()
    return JsonResponse({'ok': True})


# ── API: تراکنش (ویرایش inline سه ستون + گروهی) ───────────────────────────

@login_required
@require_finance
@require_http_methods(['PATCH'])
def tx_edit(request, pk):
    t = get_object_or_404(Transaction, pk=pk)
    d = _body(request)
    if 'project' in d:
        t.project_id = d['project'] or None
    if 'category' in d:
        t.category_id = d['category'] or None
    if 'note' in d:
        t.note = d['note']
    t.save(update_fields=['project', 'category', 'note', 'updated_at'])

    # هشدارِ نرم (بدونِ بلاک) اگر این نسبت‌دهی ناسازگاریِ حسابداری ساخت
    warning = _tx_anomaly_warning(t)
    return JsonResponse({'ok': True, 'warning': warning})


def _tx_anomaly_warning(t):
    """اگر نسبت‌دهیِ پروژه/بابت به این تراکنش ناسازگاری ساخت، پیامِ هشدار برگردان (وگرنه None)."""
    from .balances import project_balance, salary_balance
    if t.project_id:
        if project_balance(t.project_id) > 0:
            return 'مانده‌ی این پروژه مثبت شد (واریزی بیش از فاکتورها) — شاید فاکتوری ثبت نشده باشد.'
    if t.category_id and getattr(t.category, 'colleague_id', None):
        if salary_balance(t.category.colleague_id) < 0:
            return 'پرداختِ حقوق این همکار از تعهدش بیشتر شد (اضافه‌پرداخت).'
    if t.bank_account_id and t.bank_account.balance < 0:
        return f'مانده‌ی بانکِ «{t.bank_account.name}» منفی شد.'
    return None


@login_required
@require_finance
@require_http_methods(['POST'])
def tx_bulk(request):
    d = _body(request)
    ids = d.get('ids', [])
    action = d.get('action')
    qs = Transaction.objects.filter(id__in=ids)
    if action == 'set_project':
        qs.update(project_id=d.get('project') or None)
    elif action == 'set_category':
        qs.update(category_id=d.get('category') or None)
    elif action == 'set_note':
        qs.update(note=d.get('note', ''))
    elif action == 'delete':
        qs.delete()
    else:
        return JsonResponse({'detail': 'action نامعتبر'}, status=400)
    return JsonResponse({'ok': True, 'count': len(ids)})


# ── API: ورود اکسل (پیش‌نمایش + تایید) ────────────────────────────────────

def _cell(r, i):
    return r[i] if i < len(r) else None


def _num(v):
    """عددِ خامِ اکسل (float) یا رشته‌ی فارسی → int با حفظِ علامت."""
    if isinstance(v, (int, float)):
        return int(round(v))
    return parse_amount(v)


def _row_text(r):
    return ' '.join('' if c is None else str(c) for c in r)


# امضای هدرِ هر قالب — برای تشخیصِ خودکار و برای یافتنِ ردیفِ هدر وقتی قالب دستی انتخاب شده
_FMT_SIGNATURE = {
    'mehr': lambda t: 'زمان تراکنش' in t and 'مبلغ' in t,
    'tejarat': lambda t: 'موجودی حساب' in t and ('شرح تراکنش' in t or 'RRN' in t),
    'saman': lambda t: 'شرح سند' in t and 'واریز' in t,
}


def _detect_format(rows):
    """(format, header_index) با اسکنِ ۲۵ ردیفِ اول. ترتیب مهم است: هدرِ تجارت هم
    «شرح سند/واریز» دارد، پس مهر (با «زمان تراکنش») و تجارت (با «موجودی حساب») که
    یکتا هستند اول چک می‌شوند؛ سامان/موبایلت (قالبِ کلاسیک) fallback است."""
    for i, r in enumerate(rows[:25]):
        t = _row_text(r)
        if _FMT_SIGNATURE['mehr'](t):
            return 'mehr', i
        if _FMT_SIGNATURE['tejarat'](t):
            return 'tejarat', i
        if _FMT_SIGNATURE['saman'](t):
            return 'saman', i
    return 'saman', -1  # قالبِ کلاسیک بدونِ هدرِ شناخته‌شده (هدر با تاریخِ نامعتبر رد می‌شود)


def _find_header(rows, fmt):
    """ردیفِ هدرِ همان قالبِ **دستی‌انتخاب‌شده** را پیدا می‌کند؛ اگر نبود ‎-1‎ (از ابتدا)."""
    sig = _FMT_SIGNATURE.get(fmt)
    if sig:
        for i, r in enumerate(rows[:40]):
            if sig(_row_text(r)):
                return i
    return -1


# ── نگاشتِ ستون‌ها بر اساسِ نامِ هدر (مقاوم به تغییر جای ستون‌ها / قالبِ متفاوتِ هر بانک) ──
def _hmap(row):
    """{برچسبِ هدر → ایندکسِ ستون} از یک ردیفِ هدر."""
    return {str(c).strip(): i for i, c in enumerate(row) if c is not None and str(c).strip()}


def _col(hmap, *needles):
    """ایندکسِ ستونی که برچسبش یکی از needleها را دارد. سه لایه به ترتیبِ دقت: تطبیقِ
    دقیق → آغازشونده‌با → شامل. لایه‌ی «آغازشونده» ابهامِ «شناسه واریز» با «واریز (ریال)»
    را حل می‌کند (فقط دومی با «واریز» شروع می‌شود)."""
    for lbl, i in hmap.items():
        if lbl in needles:
            return i
    for lbl, i in hmap.items():
        if any(lbl.startswith(n) for n in needles):
            return i
    for lbl, i in hmap.items():
        if any(n in lbl for n in needles):
            return i
    return None


def _at(r, i):
    return r[i] if (i is not None and i < len(r)) else None


def _parse_saman(rows, h):
    """سامان (و موبایلت/کلاسیک). نگاشت با نامِ هدر — پس هم قالبِ ۷ستونیِ موبایلت و هم
    قالبِ کاملِ سامان (۱۱ ستون: ردیف/تاریخ/شرح سند/شماره سند/شماره برگه/شناسه واریز/شعبه/
    واریز/برداشت/مانده/توضیحات کاربر) کار می‌کند. توضیحات ← «توضیحات کاربر»."""
    if h < 0:
        h = 0
    hm = _hmap(rows[h])
    ci_date, ci_desc = _col(hm, 'تاریخ'), _col(hm, 'شرح سند', 'شرح')
    ci_dep, ci_wd = _col(hm, 'واریز'), _col(hm, 'برداشت')
    ci_bal, ci_note = _col(hm, 'مانده'), _col(hm, 'توضیحات')
    out = []
    for r in rows[h + 1:]:
        d, tm = parse_excel_date(_at(r, ci_date))
        if d is None:
            continue
        bal = _at(r, ci_bal)
        out.append({
            'date': d.isoformat(), 'time': tm,
            'description': str(_at(r, ci_desc) or '').strip(),
            'deposit': parse_amount(_at(r, ci_dep)),
            'withdrawal': parse_amount(_at(r, ci_wd)),
            'balance': parse_amount(bal) if bal not in (None, '') else None,
            'user_note': str(_at(r, ci_note) or '').strip(),
        })
    return out


def _parse_mehr(rows, h):
    """بانک مهر (سلول‌های مرج‌شده؛ هر تراکنش ۲ ردیفِ اکسل، ردیفِ دوم خالی). نگاشت با نامِ
    هدر: توضیح / مبلغ(علامت‌دار: +واریز −برداشت) / نوع / زمان تراکنش / مانده. ردیف‌های
    خالیِ بینابین با تاریخِ نامعتبر رد می‌شوند."""
    if h < 0:
        return []
    hm = _hmap(rows[h])
    ci_desc, ci_amt = _col(hm, 'توضیح'), _col(hm, 'مبلغ')
    ci_type, ci_dt = _col(hm, 'نوع'), _col(hm, 'زمان')
    ci_bal = _col(hm, 'مانده')
    out = []
    for r in rows[h + 1:]:
        d, tm = parse_excel_date(_at(r, ci_dt))
        if d is None:
            continue
        amount = _num(_at(r, ci_amt))
        typ = str(_at(r, ci_type) or '')
        if amount < 0 or 'برداشت' in typ:
            deposit, withdrawal = 0, abs(amount)
        else:
            deposit, withdrawal = abs(amount), 0
        bal = _at(r, ci_bal)
        out.append({
            'date': d.isoformat(), 'time': tm,
            'description': str(_at(r, ci_desc) or '').strip(),
            'deposit': deposit, 'withdrawal': withdrawal,
            'balance': _num(bal) if bal not in (None, '') else None,
            'user_note': '',
        })
    return out


def _parse_tejarat(rows, h):
    """بانک تجارت (هدر پایینِ «جزئیاتِ دوره»، ستونِ اولِ خالی). نگاشت با نامِ هدر:
    شرح سند → **توضیحات**؛ شرح تراکنش (یا شرح عملیات) → **شرح سندِ خروجی**؛
    واریز/برداشت/موجودی حساب/زمان/تاریخ."""
    if h < 0:
        return []
    hm = _hmap(rows[h])
    ci_date, ci_time = _col(hm, 'تاریخ'), _col(hm, 'زمان')
    ci_dep, ci_wd = _col(hm, 'واریز'), _col(hm, 'برداشت')
    ci_bal = _col(hm, 'موجودی')
    ci_snd = _col(hm, 'شرح سند')      # → توضیحات
    ci_trx = _col(hm, 'شرح تراکنش')   # → شرح سندِ خروجی (توضیحِ اصلی)
    ci_op = _col(hm, 'شرح عملیات')
    out = []
    for r in rows[h + 1:]:
        d, tm = parse_excel_date(f"{_at(r, ci_date)} {_at(r, ci_time)}")
        if d is None:
            continue
        desc = (str(_at(r, ci_trx) or '').strip()
                or str(_at(r, ci_op) or '').strip()
                or str(_at(r, ci_snd) or '').strip())
        bal = _at(r, ci_bal)
        out.append({
            'date': d.isoformat(), 'time': tm,
            'description': desc,
            'deposit': parse_amount(_at(r, ci_dep)),
            'withdrawal': parse_amount(_at(r, ci_wd)),
            'balance': parse_amount(bal) if bal not in (None, '') else None,
            'user_note': str(_at(r, ci_snd) or '').strip(),
        })
    return out


# «سایر» = قالبِ استانداردِ فعلی (همان پارسرِ سامان/کلاسیک)
_PARSERS = {'saman': _parse_saman, 'mehr': _parse_mehr,
            'tejarat': _parse_tejarat, 'other': _parse_saman}
_FORMAT_LABEL = {'saman': 'سامان', 'mehr': 'مهر',
                 'tejarat': 'تجارت', 'other': 'قالبِ استاندارد'}


def _parse_workbook(f, fmt=None):
    """اکسلِ بانک را می‌خواند. اگر `fmt` (مهر/سامان/تجارت/سایر) داده شود همان قالب
    استفاده می‌شود؛ وگرنه از روی ساختار **خودکار تشخیص** می‌دهد. خروجی: (format_key, rows)."""
    from openpyxl import load_workbook

    from core.jalali import to_en_digits
    wb = load_workbook(f, read_only=True, data_only=True)
    ws = wb.active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    if fmt in _PARSERS:
        real = 'saman' if fmt == 'other' else fmt
        parsed = _PARSERS[fmt](rows, _find_header(rows, real))
    else:
        fmt, h = _detect_format(rows)
        parsed = _PARSERS[fmt](rows, h)
    # اعداد را همیشه لاتین ذخیره کن (شرحِ بانکی معمولاً ارقامِ فارسی دارد)
    for r in parsed:
        r['description'] = to_en_digits(r.get('description', ''))
        r['user_note'] = to_en_digits(r.get('user_note', ''))
    return fmt, parsed


@login_required
@require_finance
@require_http_methods(['POST'])
def import_preview(request):
    import datetime as _dt
    bank_id = request.POST.get('bank')
    fmt_sel = (request.POST.get('format') or '').strip() or None  # مهر/سامان/تجارت/سایر یا None=خودکار
    f = request.FILES.get('file')
    if not bank_id or not f:
        return JsonResponse({'detail': 'حساب و فایل لازم است'}, status=400)
    try:
        fmt, rows = _parse_workbook(f, fmt_sel)
    except Exception as e:  # noqa: BLE001
        return JsonResponse({'detail': f'خطا در خواندن فایل: {e}'}, status=400)
    if not rows:
        return JsonResponse({'detail': 'ردیفِ معتبری در فایل پیدا نشد — شاید قالبِ انتخابی '
                                       'با فایل نمی‌خواند.'}, status=400)
    existing = set(Transaction.objects.filter(bank_account_id=bank_id).values_list('import_hash', flat=True))
    new, dup = 0, 0
    for row in rows:
        h = Transaction.make_hash(bank_id, row['date'], row['description'],
                                  row['deposit'], row['withdrawal'], row['balance'])
        row['hash'] = h
        row['dup'] = h in existing
        row['date_fa'] = _fa_date(row['date'])
        dup += row['dup']
        new += not row['dup']
    return JsonResponse({'rows': rows, 'new': new, 'dup': dup,
                         'format': _FORMAT_LABEL.get(fmt, fmt)})


def _fa_date(iso):
    from core.jalali import format_jalali
    import datetime as _dt
    try:
        return format_jalali(_dt.date.fromisoformat(iso))
    except Exception:  # noqa: BLE001
        return iso


@login_required
@require_finance
@require_http_methods(['POST'])
def import_confirm(request):
    import datetime as _dt
    d = _body(request)
    bank_id = d.get('bank')
    rows = d.get('rows', [])
    bank = get_object_or_404(BankAccount, pk=bank_id)
    existing = set(Transaction.objects.filter(bank_account=bank).values_list('import_hash', flat=True))
    created = 0
    for row in rows:
        h = row.get('hash')
        if not h or h in existing:
            continue
        note = row.get('user_note', '')
        Transaction.objects.create(
            bank_account=bank, date=_dt.date.fromisoformat(row['date']), time=row.get('time', ''),
            description=row.get('description', ''), deposit=row.get('deposit', 0),
            withdrawal=row.get('withdrawal', 0),
            balance=row.get('balance'), user_note=note,
            # ستون «توضیحات کاربر» اکسل مستقیم در فیلد قابل‌ویرایشِ «توضیحات» می‌نشیند
            note=note,
            import_hash=h, created_by=request.user)
        existing.add(h)
        created += 1
    return JsonResponse({'ok': True, 'created': created})


# ── API: حقوق ─────────────────────────────────────────────────────────────

@login_required
@require_finance
@require_http_methods(['POST'])
def payroll_create(request):
    d = _body(request)
    try:
        p, _ = Payroll.objects.get_or_create(
            colleague_id=d['colleague'], year=int(d['year']), month=int(d['month']))
    except (KeyError, ValueError):
        return JsonResponse({'detail': 'همکار/سال/ماه لازم است'}, status=400)
    for it in d.get('items', []):
        if it.get('title'):
            PayrollItem.objects.create(payroll=p, title=it['title'], amount=parse_amount(it.get('amount', 0)))
    return JsonResponse({'id': p.id}, status=201)


@login_required
@require_finance
@require_http_methods(['PATCH', 'DELETE'])
def payroll_edit(request, pk):
    p = get_object_or_404(Payroll, pk=pk)
    if request.method == 'DELETE':
        p.delete()
        return JsonResponse({'ok': True})
    d = _body(request)
    if 'paid_amount' in d:
        p.paid_amount = parse_amount(d['paid_amount'])
    if 'note' in d:
        p.note = d['note']
    if 'colleague' in d and d['colleague']:
        p.colleague_id = d['colleague']
    if 'year' in d:
        try:
            p.year = int(d['year'])
        except (ValueError, TypeError):
            pass
    if 'month' in d:
        try:
            p.month = int(d['month'])
        except (ValueError, TypeError):
            pass
    from django.db import IntegrityError
    try:
        p.save()
    except IntegrityError:
        return JsonResponse({'detail': 'برای این همکار در این ماه از قبل حقوق ثبت شده'}, status=400)
    if 'items' in d:
        p.items.all().delete()
        for it in d['items']:
            if it.get('title'):
                PayrollItem.objects.create(payroll=p, title=it['title'], amount=parse_amount(it.get('amount', 0)))
    return JsonResponse({'ok': True, 'total': p.total, 'remaining': p.remaining, 'status': p.status})


# ── API: فاکتور ───────────────────────────────────────────────────────────

def _parse_qty(value):
    """تعداد را با حفظِ اعشار پارس می‌کند (برخلافِ parse_amount که گرد می‌کند)."""
    from core.jalali import to_en_digits
    import re as _re
    if value in (None, ''):
        return 1
    s = _re.sub(r'[^\d.]', '', to_en_digits(str(value)))
    try:
        return round(float(s), 2) or 1
    except ValueError:
        return 1


def _save_lines(invoice, lines):
    """ردیف‌های فاکتور را از نو می‌سازد (منبع واحد = آرایه‌ی ورودی)."""
    invoice.lines.all().delete()
    for i, li in enumerate(lines):
        # ردیفِ کاملاً خالی را رد کن
        if not (li.get('description') or li.get('category') or
                parse_amount(li.get('unit_price', 0))):
            continue
        InvoiceLine.objects.create(
            invoice=invoice, order=i,
            category_id=li.get('category') or None,
            description=(li.get('description') or '')[:255],
            qty=_parse_qty(li.get('qty', 1)),
            unit_price=parse_amount(li.get('unit_price', 0)),
            tax=parse_amount(li.get('tax', 0)),
            discount=parse_amount(li.get('discount', 0)),
        )


def _invoice_totals(inv):
    return {'subtotal': inv.subtotal, 'tax_total': inv.tax_total,
            'discount_total': inv.discount_total, 'grand_total': inv.grand_total}


@login_required
@require_finance
@require_http_methods(['POST'])
def invoice_create(request):
    d = _body(request)
    issue = _pj(d.get('issue_date'))
    if not issue:
        return JsonResponse({'detail': 'تاریخ ثبت لازم است'}, status=400)
    if not (d.get('project') or None):
        return JsonResponse({'detail': 'پروژه لازم است'}, status=400)
    inv = Invoice.objects.create(
        issue_date=issue, project_id=d.get('project') or None,
        description=(d.get('description') or '').strip(),
        due_date=_pj(d.get('due_date')), created_by=request.user)
    _save_lines(inv, d.get('lines', []))
    return JsonResponse({'id': inv.id, 'number': inv.number}, status=201)


@login_required
@require_finance
@require_http_methods(['PATCH', 'DELETE'])
def invoice_edit(request, pk):
    inv = get_object_or_404(Invoice, pk=pk)
    if request.method == 'DELETE':
        inv.delete()
        return JsonResponse({'ok': True})
    d = _body(request)
    if 'issue_date' in d:
        issue = _pj(d['issue_date'])
        if issue:
            inv.issue_date = issue
    if 'project' in d:
        if not d['project']:
            return JsonResponse({'detail': 'پروژه لازم است'}, status=400)
        inv.project_id = d['project']
    if 'description' in d:
        inv.description = (d['description'] or '').strip()
    if 'due_date' in d:
        inv.due_date = _pj(d['due_date'])
    inv.save()
    if 'lines' in d:
        _save_lines(inv, d['lines'])
    return JsonResponse({'ok': True, **_invoice_totals(inv)})
